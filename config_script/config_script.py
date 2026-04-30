# $language = "Python3"
# $interface = "1.0"

import re
import os

# ====== 获取脚本所在目录 ======
SCRIPT_DIR = os.path.dirname(crt.ScriptFullName)

# ====== 文件路径 ======
TEMPLATE_FILE = os.path.join(SCRIPT_DIR, "config_template.txt")
EXCEL_FILE   = os.path.join(SCRIPT_DIR, "devices_vars.xlsx")

# ====== 导入 Excel ======
try:
    from openpyxl import load_workbook
except ImportError:
    crt.Dialog.MessageBox("需要安装 openpyxl 模块")
    raise

# ====== 输入设备名称 ======
device_name = crt.Dialog.Prompt(
    "请输入设备名称（对应Excel中的 DEVICE_NAME）:",
    "设备选择",
    ""
)

if not device_name:
    crt.Dialog.MessageBox("未输入设备名称，退出")
    exit()

# ====== 读取 Excel ======
wb = load_workbook(EXCEL_FILE)
ws = wb.active

headers = []
device_data = None

# 获取表头
for cell in ws[1]:
    headers.append(cell.value)

# 查找设备
for row in ws.iter_rows(min_row=2, values_only=True):
    row_dict = dict(zip(headers, row))

    if str(row_dict.get("DEVICE_NAME", "")).strip().lower() == device_name.strip().lower():
        device_data = row_dict
        break

if not device_data:
    crt.Dialog.MessageBox("未找到设备: {}".format(device_name))
    exit()

# ====== 读取模板 ======
with open(TEMPLATE_FILE, "r", encoding="utf-8") as f:
    config = f.read()

# ====== 提取模板变量 ======
pattern = re.compile(r"\$\{(\w+)\}")
vars_in_template = set(pattern.findall(config))

# ====== 检查变量完整性 ======
missing_vars = [v for v in vars_in_template if v not in device_data]

if missing_vars:
    crt.Dialog.MessageBox(
        "以下变量在Excel中未定义:\n" + "\n".join(missing_vars)
    )
    exit()

# ====== 变量替换 ======
def replace_var(match):
    return str(device_data.get(match.group(1), match.group(0)))

config = pattern.sub(replace_var, config)

# ====== 配置预览 ======
confirm = crt.Dialog.MessageBox(
    "即将下发以下配置:\n\n{}\n\n是否继续？".format(config),
    "配置预览",
    1  # Yes/No
)

if confirm != 1:
    exit()

# ====== 判断设备品牌 ======
# 要求 Excel 中有一列：MANUFACTURER（cisco / h3c）
manufacturer = str(device_data.get("MANUFACTURER", "")).lower()

if manufacturer == "cisco":
    enter_cmd = "configure terminal"
    exit_cmd = "end"
    prompt = "#"
elif manufacturer == "h3c":
    enter_cmd = "system-view"
    exit_cmd = "return"
    prompt = "]"
else:
    crt.Dialog.MessageBox("不支持的设备类型: {}".format(manufacturer))
    exit()

# ====== 开始下发 ======
crt.Screen.Synchronous = True

# 进入配置模式
crt.Screen.Send(enter_cmd + "\r")
crt.Screen.WaitForString(prompt)

# 逐行发送配置
for line in config.splitlines():
    if not line.strip():
        continue  # 跳过空行

    crt.Screen.Send(line + "\r")
    crt.Screen.WaitForString(prompt)

# 退出配置模式
crt.Screen.Send(exit_cmd + "\r")
crt.Screen.WaitForString(prompt)

crt.Screen.Synchronous = False

crt.Dialog.MessageBox("配置已完成！")